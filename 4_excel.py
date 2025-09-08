from openpyxl import load_workbook
from openpyxl.drawing.image import  Image


wb= load_workbook('files/gastos.xlsx')
planilha = wb['PLGastos']


valor_total=0
#1- soma de valores
for i in range(2,62):
    #print(planilha['B%s' %i].value)
    valor= float(planilha['B%s' %i].value)
    valor_total += valor
#print(valor_total)
planilha['C62']= valor_total

#wb.save(filename='files/gastos.xlsx')
#2- mesclar celula
planilha['A62']='Total'

planilha.merge_cells('A62:B62')

#wb.save(filename='files/gastos.xlsx')
#3- inserindo imagens
img= Image('files/mussum.jpg')
planilha.add_image(img, 'A63')

#wb.save(filename='files/gastos.xlsx')

#4- deletando celulas
planilha.delete_rows(1)
planilha.delete_cols(3)

wb.save(filename='files/gastos2.xlsx')

