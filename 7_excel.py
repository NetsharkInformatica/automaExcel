##Criação de graficom com Python
from openpyxl import Workbook,load_workbook
from openpyxl.chart import AreaChart,Reference,series,BarChart


dict_ano={}

#1importando despesas

arquivo1=load_workbook(filename='files/despesa.xlsx')
ws1=arquivo1['Despesas']
max_linhas=ws1.max_row


for i in range(2,max_linhas +1):
    #print(ws1['A%s' %i].value)
    dict_ano[ws1['A%s' %i].value]= {'despesa':ws1['B%s'%i].value}
    #print(dict_ano)
    
#2- Lendo o arqivo receita
arquivo2=load_workbook(filename='files/receita.xlsx')
ws2=arquivo2['Receita']
max_linhas=ws2.max_row

for i in range(2,max_linhas +1):
    #print(ws2['A%s' %j].value)
   dict_ano[ ws2['A%s'%i].value]['receita']=ws2['B%s'%i].value
   
#print(dict_ano)
#3-Criando a planilha
wb=Workbook()
ws= wb.active

ws['A1']='Ano'
ws['B1']='Despesa'
ws['C1']='Receita'

i=2

for key,value in dict_ano.items():
    ws['A%s'%i]=key
    ws['B%s'%i]=value['despesa']
    ws['C%s'%i]=value['receita']
    i +=1
    


#4- criando as planilhas

chart1=BarChart()
chart1.type='col'
chart1.style= 13
chart1.title='Receita X Despesa por ano'
chart1.x_axis.title='Ano'
chart1.y_axis.title='R$'

data=Reference(
    ws,
    min_col=2,
    max_col=3,
    min_row=1,
    max_row=i,
    
)
anos =Reference(
    ws,
    min_col=2,
    min_row=2,
    max_row=i,
    
)

# data = Reference(
#     ws,
#     min_col=2,
#     max_col=3,
#     min_row=1,  # Inclui os cabeçalhos
#     max_row=i-1  # Última linha com dados
# )

# # Categorias (anos da coluna A)
# anos = Reference(
#     ws,
#     min_col=1,
#     max_col=1,
#     min_row=2,  # Começa da linha 2 (sem o cabeçalho)
#     max_row=i-1  # Última linha com dados
# )

chart1.add_data(data,titles_from_data=True)

chart1.set_categories(anos)
chart1.shape=4

ws.add_chart(chart1,'A%s' %(i+2))

wb.save(filename='files/demonstrativo.xlsx')
    


