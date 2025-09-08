##JUNÇÂO DE PLANILHAS##
from openpyxl import load_workbook,Workbook
lista_arquivos=['gastos','gastos2']

#criar nova planilha

wb=Workbook()
nome_arquivo ='files/resultado.xlsx'
for nome in lista_arquivos:
    arquivo= load_workbook(filename='files/%s.xlsx' %nome)
    sheet=arquivo[nome]
    max_linha= sheet.max_row
    max_coluna= sheet.max_column
    ws= wb.create_sheet(title=nome)
    
    #iterar valores da planilha
    for i in range(1,max_linha + 1):
        for j in range(1,max_coluna +1):
            data= sheet.cell(row=i ,column=j)
            ws.cell(row=i,column=j).value= data.value
            
wb.remove(wb['Sheet'])
wb.save(nome_arquivo)
    
    